import io
import shutil
import tempfile
import zipfile
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
ZIP_PATH = ROOT / "school files.zip"
BACKUP_PATH = ROOT / "outputs" / "school files.before_robert_napier_withdrawals.zip"
TARGET = "RobertNapier_UPDATED.xlsx"
WITHDRAWN = {
    ("lohla", "pearce"),
    ("david", "udo"),
    ("jacob", "cave"),
}


def clean(value):
    return "" if value is None else str(value).strip().lower()


def main():
    BACKUP_PATH.parent.mkdir(parents=True, exist_ok=True)
    if not BACKUP_PATH.exists():
        shutil.copy2(ZIP_PATH, BACKUP_PATH)

    with zipfile.ZipFile(ZIP_PATH, "r") as source_zip:
        workbook_bytes = source_zip.read(TARGET)
        workbook = openpyxl.load_workbook(io.BytesIO(workbook_bytes))
        worksheet = workbook["Sheet1"]

        withdrawn_rows = []
        for row_idx in range(2, worksheet.max_row + 1):
            forename = clean(worksheet.cell(row_idx, 2).value)
            surname = clean(worksheet.cell(row_idx, 3).value)
            if (forename, surname) in WITHDRAWN:
                withdrawn_rows.append(
                    {
                        "row": row_idx,
                        "forename": str(worksheet.cell(row_idx, 2).value).strip(),
                        "surname": str(worksheet.cell(row_idx, 3).value).strip(),
                        "dob": str(worksheet.cell(row_idx, 4).value),
                    }
                )

        found = {(row["forename"].lower(), row["surname"].lower()) for row in withdrawn_rows}
        missing = WITHDRAWN - found
        if missing:
            raise RuntimeError(f"Did not find withdrawn Robert Napier pupils: {sorted(missing)}")

        for row in sorted(withdrawn_rows, key=lambda item: item["row"], reverse=True):
            worksheet.delete_rows(row["row"], 1)

        output = io.BytesIO()
        workbook.save(output)
        replacement_bytes = output.getvalue()

        with tempfile.NamedTemporaryFile(delete=False, suffix=".zip") as temp_file:
            temp_path = Path(temp_file.name)

        try:
            with zipfile.ZipFile(temp_path, "w", compression=zipfile.ZIP_DEFLATED) as out_zip:
                for item in source_zip.infolist():
                    data = replacement_bytes if item.filename == TARGET else source_zip.read(item.filename)
                    out_zip.writestr(item, data)
            shutil.move(str(temp_path), ZIP_PATH)
        finally:
            if temp_path.exists():
                temp_path.unlink()

    print("Updated", TARGET)
    print("Withdrawn rows:", withdrawn_rows)
    print("Backup:", BACKUP_PATH)


if __name__ == "__main__":
    main()
