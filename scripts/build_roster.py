import argparse
import csv
import datetime as dt
import io
import json
import re
import zipfile
from pathlib import Path

import openpyxl
from openpyxl.utils.datetime import from_excel


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_ZIP = ROOT / "school files.zip"
DEFAULT_OUTPUT = ROOT / "outputs"


HEADER_SYNONYMS = {
    "school": [
        "school",
        "school name",
        "school name no abbreviations",
    ],
    "forename": [
        "forename",
        "first name",
        "firstname",
        "given name",
        "pupil first name",
        "pupil forename",
        "student first name",
        "student forename",
        "child first name",
        "child forename",
    ],
    "surname": [
        "surname",
        "last name",
        "lastname",
        "family name",
        "pupil last name",
        "pupil surname",
        "student last name",
        "student surname",
        "child last name",
        "child surname",
    ],
    "dob": [
        "dob",
        "d o b",
        "date of birth",
        "birth date",
        "birthday",
    ],
}


DATE_FORMAT_HINTS = ("d", "m", "y", "yy", "yyyy", "date")
MIN_YEAR = 2008
MAX_YEAR = 2016
ACCEPT_UNEXPECTED_ROSTER_COUNTS = True
ACCEPT_MISSING_DOBS = True
ACCEPT_DEFAULTED_DOB_REPAIRS = True
MANUAL_DOB_OVERRIDES = {
    ("Pupil Data Submission Outwood Academy Freeston (1).xlsx", 23): {
        "dob": dt.date(2012, 10, 4),
        "note": "Manual correction: raw year 0212 is a typo for 2012; parsed using school-level DMY convention.",
    },
}
MANUAL_EXCLUDED_ROSTER_ROWS = {
    ("Commando Joe's Students - Liberty Academy (1).xlsx", 23): "Confirmed duplicate of Scarlett Conkerton row 5.",
    ("DonValley.xlsx", 23): "Confirmed duplicate of Ebony Tilley row 17.",
}
MONTHS = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


def norm_text(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"[_\n\r\t]+", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def norm_name(value):
    text = norm_text(value)
    return re.sub(r"[^a-z]", "", text)


def compact_header(value):
    return norm_text(value).replace(" ", "")


def header_match(cell_value, field):
    cell = norm_text(cell_value)
    compact = cell.replace(" ", "")
    if not cell:
        return False
    for synonym in HEADER_SYNONYMS[field]:
        syn = norm_text(synonym)
        syn_compact = syn.replace(" ", "")
        if cell == syn or compact == syn_compact:
            return True
        if syn and syn in cell:
            return True
        if syn_compact and syn_compact in compact:
            return True
    return False


def likely_date_format(cell):
    fmt = str(cell.number_format or "").lower()
    return any(token in fmt for token in DATE_FORMAT_HINTS)


def clean_school_from_filename(filename):
    name = Path(filename).stem
    name = re.sub(r"\(\d+\)", "", name)
    name = re.sub(r"\bupdated\b", "", name, flags=re.I)
    name = re.sub(r"\bcopy of\b", "", name, flags=re.I)
    name = re.sub(r"\bpupil data submission template\b", "", name, flags=re.I)
    name = re.sub(r"\bpupil data submission\b", "", name, flags=re.I)
    name = re.sub(r"\bpupil submission data\b", "", name, flags=re.I)
    name = re.sub(r"\bstudent list\b", "", name, flags=re.I)
    name = re.sub(r"\bcommando joe'?s?\b", "", name, flags=re.I)
    name = re.sub(r"\bcommando\b", "", name, flags=re.I)
    name = re.sub(r"\bcj\b", "", name, flags=re.I)
    name = re.sub(r"\bno password\b", "", name, flags=re.I)
    name = re.sub(r"[_\-]+", " ", name)
    return re.sub(r"\s+", " ", name).strip(" -_")


def find_header(ws):
    best = None
    max_row = min(ws.max_row or 1, 60)
    max_col = min(ws.max_column or 1, 40)
    for row_idx in range(1, max_row + 1):
        field_cols = {}
        row_values = [ws.cell(row_idx, col_idx).value for col_idx in range(1, max_col + 1)]
        for col_idx, value in enumerate(row_values, start=1):
            for field in HEADER_SYNONYMS:
                if field not in field_cols and header_match(value, field):
                    field_cols[field] = col_idx
        score = sum(1 for f in ("forename", "surname", "dob") if f in field_cols)
        if score:
            candidate = {
                "score": score,
                "row": row_idx,
                "columns": field_cols,
                "row_values": [str(v) if v is not None else "" for v in row_values],
            }
            if best is None or candidate["score"] > best["score"]:
                best = candidate
    return best


def count_row_values(ws, row_idx, columns):
    return sum(1 for col_idx in columns if ws.cell(row_idx, col_idx).value not in (None, ""))


def extract_rows(ws, header):
    fields = header["columns"]
    relevant_cols = [fields[field] for field in ("forename", "surname", "dob") if field in fields]
    rows = []
    blank_run = 0
    for row_idx in range(header["row"] + 1, (ws.max_row or 1) + 1):
        non_empty = count_row_values(ws, row_idx, relevant_cols)
        if non_empty == 0:
            blank_run += 1
            if blank_run >= 8 and rows:
                break
            continue
        blank_run = 0
        row = {"source_row": row_idx}
        for field, col_idx in fields.items():
            cell = ws.cell(row_idx, col_idx)
            row[field] = cell.value
            row[f"{field}_number_format"] = cell.number_format
            row[f"{field}_is_date"] = bool(cell.is_date)
        if row.get("forename") or row.get("surname") or row.get("dob"):
            rows.append(row)
    return rows


def parse_text_date_tokens(text):
    raw = str(text).strip()
    if not raw:
        return None

    named_month = re.fullmatch(
        r"(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+)\s+(\d{2,5})",
        raw,
        flags=re.I,
    )
    if named_month:
        day_text, month_text, year_text = named_month.groups()
        month = MONTHS.get(month_text.lower())
        if not month:
            return None
        year = repair_year(int(year_text))
        return {"kind": "named_month", "year": year, "month": month, "day": int(day_text), "raw": raw}

    iso = re.fullmatch(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", raw)
    if iso:
        y, m, d = map(int, iso.groups())
        return {"kind": "iso", "year": y, "a": d, "b": m, "raw": raw}

    compact = re.fullmatch(r"(\d{1,2})[./\-](\d{2})(\d{4,5})", raw)
    if compact:
        a, b, y = compact.groups()
        return {"kind": "ambiguous_order", "year": repair_year(int(y)), "a": int(a), "b": int(b), "raw": raw}

    match = re.fullmatch(r"(\d{1,2})(?:st|nd|rd|th)?[./\-\s](\d{1,2})[./\-\s](\d{2,5})", raw, flags=re.I)
    if not match:
        return None
    a, b, y = map(int, match.groups())
    y = repair_year(y)
    return {"kind": "ambiguous_order", "year": y, "a": a, "b": b, "raw": raw}


def repair_year(year):
    if year < 100:
        return year + 2000 if year < 50 else year + 1900
    if 200 <= year <= 299:
        candidate = year + 1800
        if MIN_YEAR <= candidate <= MAX_YEAR:
            return candidate
    if 20000 <= year <= 29999:
        candidate = int(str(year)[0] + str(year)[2:])
        if MIN_YEAR <= candidate <= MAX_YEAR:
            return candidate
    return year


def valid_date(year, month, day):
    try:
        parsed = dt.date(year, month, day)
    except ValueError:
        return None
    if MIN_YEAR <= parsed.year <= MAX_YEAR:
        return parsed
    return None


def collect_date_evidence(raw_rows):
    evidence = {
        "excel_dates": 0,
        "text_dates": 0,
        "dmy_proofs": [],
        "mdy_proofs": [],
        "ambiguous_text": [],
        "invalid_text": [],
        "numeric_date_serials": 0,
    }
    for row in raw_rows:
        raw = row.get("dob")
        if raw in (None, ""):
            continue
        if isinstance(raw, dt.datetime):
            evidence["excel_dates"] += 1
            continue
        if isinstance(raw, dt.date):
            evidence["excel_dates"] += 1
            continue
        if isinstance(raw, (int, float)) and row.get("dob_is_date"):
            evidence["numeric_date_serials"] += 1
            continue
        tokens = parse_text_date_tokens(raw)
        if not tokens:
            evidence["invalid_text"].append(str(raw))
            continue
        evidence["text_dates"] += 1
        if tokens["kind"] == "iso":
            evidence["dmy_proofs"].append(tokens["raw"])
            continue
        if tokens["kind"] == "named_month":
            evidence["dmy_proofs"].append(tokens["raw"])
            continue
        a, b = tokens["a"], tokens["b"]
        if a > 12 and b <= 12:
            evidence["dmy_proofs"].append(tokens["raw"])
        elif b > 12 and a <= 12:
            evidence["mdy_proofs"].append(tokens["raw"])
        elif a > 12 and b > 12:
            evidence["invalid_text"].append(tokens["raw"])
        else:
            evidence["ambiguous_text"].append(tokens["raw"])
    return evidence


def infer_school_date_format(evidence):
    has_dmy = bool(evidence["dmy_proofs"])
    has_mdy = bool(evidence["mdy_proofs"])
    has_text = evidence["text_dates"] > 0
    has_excel = evidence["excel_dates"] + evidence["numeric_date_serials"] > 0
    if has_dmy and has_mdy:
        return "mixed_text_evidence"
    if has_dmy:
        return "DMY"
    if has_mdy:
        return "MDY"
    if has_text:
        return "DMY_UNPROVEN_DEFAULT"
    if has_excel:
        return "EXCEL"
    return "NO_DATES"


def parse_dob(raw, row, convention, workbook_epoch):
    if raw in (None, ""):
        return None, "missing", "no DOB value"
    if isinstance(raw, dt.datetime):
        return raw.date(), "ok", "native Excel datetime"
    if isinstance(raw, dt.date):
        return raw, "ok", "native Excel date"
    if isinstance(raw, (int, float)) and row.get("dob_is_date"):
        try:
            return from_excel(raw, epoch=workbook_epoch).date(), "ok", "Excel serial date"
        except Exception as exc:
            return None, "invalid", f"Excel serial parse failed: {exc}"

    tokens = parse_text_date_tokens(raw)
    if not tokens:
        return None, "invalid", "text did not match supported date patterns"
    if tokens["kind"] == "iso":
        parsed = valid_date(tokens["year"], tokens["b"], tokens["a"])
        return parsed, "ok" if parsed else "invalid", "ISO-like yyyy-mm-dd text"
    if tokens["kind"] == "named_month":
        parsed = valid_date(tokens["year"], tokens["month"], tokens["day"])
        note = "text parsed as day month-name year"
        if str(tokens["year"]) not in str(tokens["raw"]):
            note += "; repaired malformed year"
        return parsed, "ok" if parsed else "invalid", note

    a, b, year = tokens["a"], tokens["b"], tokens["year"]
    if convention in ("DMY", "DMY_UNPROVEN_DEFAULT"):
        day, month = a, b
    elif convention == "MDY":
        month, day = a, b
    else:
        return None, "ambiguous", f"school-level date convention is {convention}"
    parsed = valid_date(year, month, day)
    if not parsed:
        return None, "invalid", f"invalid under {convention}"
    status = "ok" if convention != "DMY_UNPROVEN_DEFAULT" else "ok_defaulted"
    return parsed, status, f"text parsed as {convention}"


def choose_best_sheet(workbook):
    candidates = []
    for ws in workbook.worksheets:
        header = find_header(ws)
        if not header:
            continue
        rows = extract_rows(ws, header)
        candidates.append((header["score"], len(rows), ws, header, rows))
    if not candidates:
        return None, None, []
    candidates.sort(key=lambda item: (item[0], item[1]), reverse=True)
    _, _, ws, header, rows = candidates[0]
    return ws, header, rows


def process_workbook(zip_file, member):
    data = zip_file.read(member)
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws, header, rows = choose_best_sheet(wb)
    school_guess = clean_school_from_filename(member)
    result = {
        "source_file": member,
        "school_from_filename": school_guess,
        "status": "ok",
        "sheet": ws.title if ws else None,
        "header": header,
        "row_count": 0,
        "date_evidence": None,
        "date_convention": None,
        "issues": [],
        "excluded_rows": [],
        "rows": [],
    }
    if not ws:
        result["status"] = "needs_decision"
        result["issues"].append("No sheet with forename/surname/DOB headers was detected.")
        return result
    if header["score"] < 3:
        result["status"] = "needs_decision"
        result["issues"].append("Detected header row is missing at least one required column.")

    evidence = collect_date_evidence(rows)
    convention = infer_school_date_format(evidence)
    result["date_evidence"] = evidence
    result["date_convention"] = convention
    if convention == "mixed_text_evidence":
        result["status"] = "needs_decision"
        result["issues"].append("DOB text values contain both DMY and MDY evidence.")
    if evidence["invalid_text"]:
        result["issues"].append("Some DOB values could not be parsed as dates.")

    roster_rows = []
    for idx, row in enumerate(rows, start=1):
        exclusion_reason = MANUAL_EXCLUDED_ROSTER_ROWS.get((member, row["source_row"]))
        if exclusion_reason:
            result["excluded_rows"].append(
                {
                    "source_row": row["source_row"],
                    "forename_raw": "" if row.get("forename") is None else str(row.get("forename")).strip(),
                    "surname_raw": "" if row.get("surname") is None else str(row.get("surname")).strip(),
                    "dob_raw": "" if row.get("dob") is None else str(row.get("dob")).strip(),
                    "reason": exclusion_reason,
                }
            )
            continue
        dob, dob_status, dob_note = parse_dob(row.get("dob"), row, convention, wb.epoch)
        override = MANUAL_DOB_OVERRIDES.get((member, row["source_row"]))
        if override:
            dob = override["dob"]
            dob_status = "ok"
            dob_note = override["note"]
        if dob_status in ("invalid", "ambiguous"):
            result["status"] = "needs_decision"
        roster_rows.append(
            {
                "roster_file": member,
                "school_from_filename": school_guess,
                "school_raw": "" if row.get("school") is None else str(row.get("school")).strip(),
                "worksheet": ws.title,
                "source_row": row["source_row"],
                "within_file_index": idx,
                "forename_raw": "" if row.get("forename") is None else str(row.get("forename")).strip(),
                "surname_raw": "" if row.get("surname") is None else str(row.get("surname")).strip(),
                "dob_raw": "" if row.get("dob") is None else str(row.get("dob")).strip(),
                "forename_norm": norm_name(row.get("forename")),
                "surname_norm": norm_name(row.get("surname")),
                "dob_iso": dob.isoformat() if dob else "",
                "birth_month": dob.month if dob else "",
                "birth_year": dob.year if dob else "",
                "dob_parse_status": dob_status,
                "dob_parse_note": dob_note,
                "school_date_convention": convention,
            }
        )
    result["rows"] = roster_rows
    result["row_count"] = len(roster_rows)
    if result["row_count"] != 24:
        result["issues"].append(f"Expected 24 pupil rows but extracted {result['row_count']}.")
    return result


def write_csv(path, rows, fieldnames):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--zip", default=str(DEFAULT_ZIP))
    parser.add_argument("--out", default=str(DEFAULT_OUTPUT))
    args = parser.parse_args()

    out_dir = Path(args.out)
    school_results = []
    roster_rows = []
    with zipfile.ZipFile(args.zip) as zf:
        members = [name for name in zf.namelist() if name.lower().endswith(".xlsx") and not Path(name).name.startswith("~$")]
        for member in members:
            try:
                result = process_workbook(zf, member)
            except Exception as exc:
                result = {
                    "source_file": member,
                    "school_from_filename": clean_school_from_filename(member),
                    "status": "needs_decision",
                    "sheet": None,
                    "header": None,
                    "row_count": 0,
                    "date_evidence": None,
                    "date_convention": None,
                    "issues": [f"Workbook processing failed: {exc}"],
                    "rows": [],
                }
            school_results.append(result)
            roster_rows.extend(result.get("rows", []))

    summary_rows = []
    for result in school_results:
        evidence = result.get("date_evidence") or {}
        summary_rows.append(
            {
                "source_file": result["source_file"],
                "school_from_filename": result["school_from_filename"],
                "status": result["status"],
                "worksheet": result.get("sheet") or "",
                "row_count": result["row_count"],
                "date_convention": result.get("date_convention") or "",
                "excel_dates": evidence.get("excel_dates", ""),
                "text_dates": evidence.get("text_dates", ""),
                "dmy_proof_count": len(evidence.get("dmy_proofs", [])) if evidence else "",
                "mdy_proof_count": len(evidence.get("mdy_proofs", [])) if evidence else "",
                "ambiguous_text_count": len(evidence.get("ambiguous_text", [])) if evidence else "",
                "invalid_text_count": len(evidence.get("invalid_text", [])) if evidence else "",
                "issues": " | ".join(result.get("issues", [])),
            }
        )

    fieldnames = [
        "roster_file",
        "school_from_filename",
        "school_raw",
        "worksheet",
        "source_row",
        "within_file_index",
        "forename_raw",
        "surname_raw",
        "dob_raw",
        "forename_norm",
        "surname_norm",
        "dob_iso",
        "birth_month",
        "birth_year",
        "dob_parse_status",
        "dob_parse_note",
        "school_date_convention",
    ]
    write_csv(out_dir / "normalized_roster.csv", roster_rows, fieldnames)
    write_csv(
        out_dir / "roster_school_summary.csv",
        summary_rows,
        [
            "source_file",
            "school_from_filename",
            "status",
            "worksheet",
            "row_count",
            "date_convention",
            "excel_dates",
            "text_dates",
            "dmy_proof_count",
            "mdy_proof_count",
            "ambiguous_text_count",
            "invalid_text_count",
            "issues",
        ],
    )
    decisions = []
    warnings = []
    for result in school_results:
        evidence = result.get("date_evidence") or {}
        if result["status"] == "needs_decision":
            decisions.append(
                {
                    "decision_type": "extract_or_parse_failure",
                    "decision_needed": "Tell the extractor how this school file should be handled.",
                    "source_file": result["source_file"],
                    "school_from_filename": result["school_from_filename"],
                    "status": result["status"],
                    "row_count": result["row_count"],
                    "date_convention": result.get("date_convention"),
                    "evidence": {
                        "issues": result.get("issues", []),
                        "dmy_proofs_sample": evidence.get("dmy_proofs", [])[:5],
                        "mdy_proofs_sample": evidence.get("mdy_proofs", [])[:5],
                        "ambiguous_text_sample": evidence.get("ambiguous_text", [])[:5],
                        "invalid_text_sample": evidence.get("invalid_text", [])[:5],
                    },
                }
            )
        if result["row_count"] != 24:
            item = (
                {
                    "decision_type": "unexpected_roster_count",
                    "decision_needed": "Decide whether to keep every extracted pupil for this school or remove/add rows to force the expected 24-pupil roster.",
                    "decision_applied": "Accepted as-is; do not force roster to 24 pupils.",
                    "source_file": result["source_file"],
                    "school_from_filename": result["school_from_filename"],
                    "row_count": result["row_count"],
                    "date_convention": result.get("date_convention"),
                    "evidence": {
                        "issues": result.get("issues", []),
                        "extracted_rows": [
                            {
                                "source_row": row["source_row"],
                                "school_raw": row["school_raw"],
                                "forename_raw": row["forename_raw"],
                                "surname_raw": row["surname_raw"],
                                "dob_raw": row["dob_raw"],
                                "dob_iso": row["dob_iso"],
                                "dob_parse_status": row["dob_parse_status"],
                            }
                            for row in result.get("rows", [])
                        ],
                    },
                }
            )
            if ACCEPT_UNEXPECTED_ROSTER_COUNTS:
                warnings.append(item)
            else:
                decisions.append(item)
        for excluded in result.get("excluded_rows", []):
            warnings.append(
                {
                    "decision_type": "manual_duplicate_exclusion",
                    "decision_applied": "Excluded confirmed duplicate pupil row from normalized roster.",
                    "source_file": result["source_file"],
                    "school_from_filename": result["school_from_filename"],
                    "source_row": excluded["source_row"],
                    "pupil": f"{excluded['forename_raw']} {excluded['surname_raw']}".strip(),
                    "evidence": {
                        "dob_raw": excluded["dob_raw"],
                        "reason": excluded["reason"],
                    },
                }
            )

    for row in roster_rows:
        if row["dob_parse_status"] == "missing":
            item = (
                {
                    "decision_type": "missing_dob",
                    "decision_needed": "Provide the pupil DOB, or confirm this pupil should remain in the roster with missing DOB and reduced matchability.",
                    "decision_applied": "Accepted in roster with missing DOB.",
                    "source_file": row["roster_file"],
                    "school": row["school_raw"] or row["school_from_filename"],
                    "source_row": row["source_row"],
                    "pupil": f"{row['forename_raw']} {row['surname_raw']}".strip(),
                    "evidence": {
                        "dob_raw": row["dob_raw"],
                        "dob_parse_status": row["dob_parse_status"],
                        "dob_parse_note": row["dob_parse_note"],
                    },
                }
            )
            if ACCEPT_MISSING_DOBS:
                warnings.append(item)
            else:
                decisions.append(item)
        elif row["dob_parse_status"] == "ok_defaulted":
            item = (
                {
                    "decision_type": "defaulted_date_format_or_repaired_year",
                    "decision_needed": "Confirm the parsed DOB is acceptable, or provide the correct DOB.",
                    "decision_applied": "Accepted repaired/defaulted DOB using the school-level date convention.",
                    "source_file": row["roster_file"],
                    "school": row["school_raw"] or row["school_from_filename"],
                    "source_row": row["source_row"],
                    "pupil": f"{row['forename_raw']} {row['surname_raw']}".strip(),
                    "evidence": {
                        "dob_raw": row["dob_raw"],
                        "dob_iso": row["dob_iso"],
                        "school_date_convention": row["school_date_convention"],
                        "dob_parse_note": row["dob_parse_note"],
                    },
                }
            )
            if ACCEPT_DEFAULTED_DOB_REPAIRS:
                warnings.append(item)
            else:
                decisions.append(item)
    (out_dir / "roster_decisions_needed.json").write_text(json.dumps(decisions, indent=2), encoding="utf-8")
    (out_dir / "roster_warnings.json").write_text(json.dumps(warnings, indent=2), encoding="utf-8")
    print(json.dumps({
        "school_files": len(school_results),
        "roster_rows": len(roster_rows),
        "needs_decision": sum(1 for row in school_results if row["status"] == "needs_decision"),
        "with_any_issue": sum(1 for row in school_results if row.get("issues")),
        "outputs": [
            str(out_dir / "normalized_roster.csv"),
            str(out_dir / "roster_school_summary.csv"),
            str(out_dir / "roster_decisions_needed.json"),
            str(out_dir / "roster_warnings.json"),
        ],
    }, indent=2))


if __name__ == "__main__":
    main()
