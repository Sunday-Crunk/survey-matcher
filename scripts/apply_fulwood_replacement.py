import datetime as dt
import io
import shutil
import tempfile
import zipfile
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
ZIP_PATH = ROOT / "school files.zip"
BACKUP_PATH = ROOT / "outputs" / "school files.before_fulwood_replacement.zip"
TARGET = "Pupil Data Submission Fulwood Academy (3).xlsx"
TARGET_ROW = 20


def main():
    BACKUP_PATH.parent.mkdir(parents=True, exist_ok=True)
    if not BACKUP_PATH.exists():
        shutil.copy2(ZIP_PATH, BACKUP_PATH)

    with zipfile.ZipFile(ZIP_PATH, "r") as source_zip:
        workbook_bytes = source_zip.read(TARGET)
        workbook = openpyxl.load_workbook(io.BytesIO(workbook_bytes))
        worksheet = workbook["Pupil data"]

        before = [worksheet.cell(TARGET_ROW, col).value for col in range(1, 7)]
        worksheet.cell(TARGET_ROW, 1).value = "Fulwood Academy"
        worksheet.cell(TARGET_ROW, 2).value = "Isaac"
        worksheet.cell(TARGET_ROW, 3).value = "Murphy"
        worksheet.cell(TARGET_ROW, 4).value = dt.datetime(2013, 2, 20)
        worksheet.cell(TARGET_ROW, 4).number_format = "DD/MM/YYYY"
        worksheet.cell(TARGET_ROW, 5).value = "Male"
        worksheet.cell(TARGET_ROW, 6).value = "P888281817035"

        output = io.BytesIO()
        workbook.save(output)
        replacement_bytes = output.getvalue()

        with tempfile.NamedTemporaryFile(delete=False, suffix=".zip") as temp_file:
            temp_path = Path(temp_file.name)

        try:
            with zipfile.ZipFile(temp_path, "w", compression=zipfile.ZIP_DEFLATED) as out_zip:
                for item in source_zip.infolist():
                    data = replacement_bytes if item.filename == TARGET else source_zip.read(item.filename)
                    out_zip.writestr(item, data)
            shutil.move(str(temp_path), ZIP_PATH)
        finally:
            if temp_path.exists():
                temp_path.unlink()

    print("Updated", TARGET, "row", TARGET_ROW)
    print("Before:", before)
    print("After:", ["Fulwood Academy", "Isaac", "Murphy", "2013-02-20", "Male", "P888281817035"])
    print("Backup:", BACKUP_PATH)


if __name__ == "__main__":
    main()
