import io
import sys
import zipfile
from pathlib import Path

import openpyxl


zip_path = Path("school files.zip")
limit = int(sys.argv[1]) if len(sys.argv) > 1 else 5

with zipfile.ZipFile(zip_path) as zf:
    names = [name for name in zf.namelist() if name.lower().endswith(".xlsx")]
    for name in names[:limit]:
        print("=" * 80)
        print(name)
        wb = openpyxl.load_workbook(io.BytesIO(zf.read(name)), data_only=True)
        print("sheets:", wb.sheetnames)
        for ws in wb.worksheets[:2]:
            print("--", ws.title, ws.max_row, ws.max_column)
            shown = 0
            for row_idx in range(1, min(ws.max_row or 1, 35) + 1):
                values = [ws.cell(row_idx, col_idx).value for col_idx in range(1, min(ws.max_column or 1, 14) + 1)]
                if any(value not in (None, "") for value in values):
                    print(row_idx, values)
                    shown += 1
                if shown >= 18:
                    break
