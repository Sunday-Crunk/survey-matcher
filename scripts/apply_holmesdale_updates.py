import io
import shutil
import tempfile
import zipfile
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
ZIP_PATH = ROOT / "school files.zip"
BACKUP_PATH = ROOT / "outputs" / "school files.before_holmesdale_updates.zip"
TARGET = "Pupil Data Submission Template - The Holmesdale School (1).xlsx"
WITHDRAWN = {
    ("isla", "dampier"),
    ("dolcie", "glazebrook"),
    ("christopher", "mills"),
}


def clean(value):
    return "" if value is None else str(value).strip().lower()


def main():
    BACKUP_PATH.parent.mkdir(parents=True, exist_ok=True)
    if not BACKUP_PATH.exists():
        shutil.copy2(ZIP_PATH, BACKUP_PATH)

    with zipfile.ZipFile(ZIP_PATH, "r") as source_zip:
        workbook_bytes = source_zip.read(TARGET)
        workbook = openpyxl.load_workbook(io.BytesIO(workbook_bytes))
        worksheet = workbook["Pupil data"]

        withdrawn_rows = []
        data_rows = []
        for row_idx in range(2, worksheet.max_row + 1):
            school = worksheet.cell(row_idx, 1).value
            col_b = worksheet.cell(row_idx, 2).value
            col_c = worksheet.cell(row_idx, 3).value
            dob = worksheet.cell(row_idx, 4).value
            if not any(value not in (None, "") for value in (school, col_b, col_c, dob)):
                continue

            current_forename = clean(col_c)
            current_surname = clean(col_b)
            if (current_forename, current_surname) in WITHDRAWN:
                withdrawn_rows.append(
                    {
                        "row": row_idx,
                        "forename": str(col_c).strip(),
                        "surname": str(col_b).strip(),
                        "dob": str(dob),
                    }
                )
            else:
                data_rows.append(row_idx)

        found = {(row["forename"].lower(), row["surname"].lower()) for row in withdrawn_rows}
        missing = WITHDRAWN - found
        if missing:
            raise RuntimeError(f"Did not find withdrawn Holmesdale pupils: {sorted(missing)}")

        for row_idx in data_rows:
            first_name_cell = worksheet.cell(row_idx, 2)
            last_name_cell = worksheet.cell(row_idx, 3)
            first_name_cell.value, last_name_cell.value = last_name_cell.value, first_name_cell.value

        for row in sorted(withdrawn_rows, key=lambda item: item["row"], reverse=True):
            worksheet.delete_rows(row["row"], 1)

        output = io.BytesIO()
        workbook.save(output)
        replacement_bytes = output.getvalue()

        with tempfile.NamedTemporaryFile(delete=False, suffix=".zip") as temp_file:
            temp_path = Path(temp_file.name)

        try:
            with zipfile.ZipFile(temp_path, "w", compression=zipfile.ZIP_DEFLATED) as out_zip:
                for item in source_zip.infolist():
                    data = replacement_bytes if item.filename == TARGET else source_zip.read(item.filename)
                    out_zip.writestr(item, data)
            shutil.move(str(temp_path), ZIP_PATH)
        finally:
            if temp_path.exists():
                temp_path.unlink()

    print("Updated", TARGET)
    print("Withdrawn rows:", withdrawn_rows)
    print("Swapped first/last name columns for", len(data_rows), "remaining pupil rows")
    print("Backup:", BACKUP_PATH)


if __name__ == "__main__":
    main()
